"""
evaluation.py
-------------
Batch experiment runner over the `ga_*_conf.py` drivers.

For every driver it runs the GA once per (target policy, seed set) pair and
collects, per run:

    target policy | seed policy | fitness | fitness primary | fitness cosmetic
                  | discovered policy

Those six columns are written to one Excel sheet per driver, and the
best-fitness-per-generation curve of every run is drawn into one plot per
driver.

    fitness = fitness primary - fitness cosmetic

(the cosmetic column is the *penalty*, always >= 0, exactly as
`ga_fitness.fitness_breakdown` reports it).

Everything a re-run would want to change lives in the RUN SETTINGS block below,
and the (target, seed) variants live in the `_spec_*` builders further down --
one builder per driver, each importing only its own module.  Nothing else in
the file needs editing to change an experiment.

Usage
-----
    python evaluation.py                 # every driver
    python evaluation.py batching        # a subset, by key
    python evaluation.py --list          # keys and the rows they would run
    python evaluation.py --plot-only     # re-plot from a previous curves.csv

Outputs land in OUT_DIR:
    evaluation.xlsx         one sheet per driver + settings + diagnostics
    curves.csv              long-format fitness curves (re-plottable)
    <driver>.png            best fitness over generations, one line per row
"""

from __future__ import annotations

import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Callable, Sequence

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd

from ga_fitness import cosmetic_penalty
import symbolic_regression as sr


# ═════════════════════════════════════════════════════════════════════════════
# RUN SETTINGS -- change these, then re-run
# ═════════════════════════════════════════════════════════════════════════════

# Random seed handed to `sr.run_ga`.  Each row is seeded identically, so rows
# differ only in their target/seed policies.  Put several values here to repeat
# every row under different GA seeds (the table then carries one row per repeat).
GA_SEEDS = (42, 7, 13, 101, 2024)

# None = use the value the driver declares (60 x 300, from ga_shared.py).  Set
# an int to override it for every driver -- useful for a quick run (e.g.
# POP_SIZE = 20, N_GENERATIONS = 10) before committing to the full sweep.
POP_SIZE      = None
N_GENERATIONS = None
N_ROLLOUTS    = None     # rollouts per fitness evaluation
HORIZON       = None     # simulation length per rollout

# How targets and seed sets are combined into rows:
#   'zip'   -> target[i] with seed_set[i]
#   'cross' -> every target with every seed set  (1 target x 2 seed sets = 2)
PAIRING = 'cross'

# Which drivers to run.  Keys are defined in SPEC_BUILDERS at the bottom.
EXPERIMENTS = ('batching', 'choice_sl', 'ratelimit', 'priority')

# Print each driver's own per-generation GA log (very verbose).
VERBOSE_GA = False

OUT_DIR   = Path(__file__).resolve().parent / "evaluation_results"
XLSX_NAME = "evaluation.xlsx"

# ── plot styling (data-viz reference palette, light surface) ─────────────────
SERIES_COLORS  = ["#2a78d6", "#eb6834", "#1baf7a", "#eda100",
                  "#e87ba4", "#008300", "#4a3aa7", "#e34948"]
SURFACE        = "#fcfcfb"
TEXT_PRIMARY   = "#0b0b0b"
TEXT_SECONDARY = "#52514e"
GRID_COLOR     = "#e4e3df"
FIG_SIZE       = (9.0, 5.6)
FIG_DPI        = 200


# Two colors, used everywhere a convergence plot needs to tell no-pre-knowledge
# apart from some-pre-knowledge -- the one distinction these plots exist to
# show.  Fixed across all four processes so the color itself carries the
# meaning, the same way it does across the rest of the figure set.
NO_PREKNOWLEDGE_COLOR   = SERIES_COLORS[0]   # blue
SOME_PREKNOWLEDGE_COLOR = SERIES_COLORS[1]   # orange


def _clean_title(title: str) -> str:
    """Category and process description only -- drop a trailing "(pi_x)"."""
    import re
    return re.sub(r"\s*\([^)]*\)\s*$", "", title)


def _preknowledge_color(seeds_label: str) -> str:
    return (NO_PREKNOWLEDGE_COLOR if "no pre-knowledge" in seeds_label.lower()
            else SOME_PREKNOWLEDGE_COLOR)


def _preknowledge_legend(ax) -> None:
    """Exactly two entries: the one distinction these plots need to show."""
    handles = [
        plt.Line2D([], [], color=NO_PREKNOWLEDGE_COLOR, linewidth=2.0),
        plt.Line2D([], [], color=SOME_PREKNOWLEDGE_COLOR, linewidth=2.0),
    ]
    _legend_below(ax, handles, ["No pre-knowledge", "Some pre-knowledge"])


def _style_axes(ax, title: str) -> None:
    """Shared chart chrome: recessive grid and axes, integer generations."""
    from matplotlib.ticker import MaxNLocator

    ax.set_title(f"{_clean_title(title)}\nbest fitness over generations",
                 color=TEXT_PRIMARY, fontsize=12, loc="left", pad=12)
    ax.set_xlabel("generation", color=TEXT_SECONDARY, fontsize=10)
    ax.set_ylabel("best fitness so far", color=TEXT_SECONDARY, fontsize=10)
    ax.xaxis.set_major_locator(MaxNLocator(integer=True))
    ax.grid(True, color=GRID_COLOR, linewidth=0.8)
    ax.set_axisbelow(True)
    for spine in ("top", "right"):
        ax.spines[spine].set_visible(False)
    for spine in ("left", "bottom"):
        ax.spines[spine].set_color(GRID_COLOR)
    ax.tick_params(colors=TEXT_SECONDARY, labelsize=9)
    ax.margins(x=0.06)


def _legend_below(ax, handles=None, labels=None) -> None:
    """Legend under the axes -- the curves rise into every in-axes corner."""
    if handles is None:
        handles, labels = ax.get_legend_handles_labels()
    legend = ax.legend(handles, labels, loc="upper left",
                       bbox_to_anchor=(0.0, -0.14), fontsize=8, frameon=False,
                       borderaxespad=0.0, handlelength=2.4, labelspacing=0.5)
    for text in legend.get_texts():
        text.set_color(TEXT_PRIMARY)


# ═════════════════════════════════════════════════════════════════════════════
# Experiment description
# ═════════════════════════════════════════════════════════════════════════════

@dataclass(frozen=True)
class Variant:
    """One named policy tree -- either a target or a single seed."""
    short: str          # plot label, e.g. 'T1'
    label: str          # human description
    tree: object


@dataclass(frozen=True)
class SeedSet:
    """The seed policies a GA run starts from.

    A *set* rather than a single tree because these experiments are built
    around crossover assembling a reference from complementary seeds -- see the
    driver docstrings.  All of them are reported in the 'seed policy' column.
    """
    short: str
    label: str
    trees: Sequence[object]


@dataclass
class ExperimentSpec:
    key: str
    module: object
    title: str                      # plot title / sheet caption
    targets: Sequence[Variant]
    seed_sets: Sequence[SeedSet]
    # counts(module, tree, reference) -> (divergences, decisions) | None
    counts: Callable
    # primary(module, counts) -> float | None
    primary: Callable

    @property
    def sheet(self) -> str:
        return self.module.__name__[:31]


@dataclass
class RowResult:
    key: str
    target: Variant
    seed_set: SeedSet
    ga_seed: int
    fitness: float
    primary: float
    cosmetic: float                                    # raw penalty, scale 1.0
    discovered: object
    cosmetic_scale: float = 1.0                        # final-generation scale
    curve: list[float] = field(default_factory=list)   # best-so-far per generation
    target_fitness: float = float('nan')
    seed_fitness: float = float('nan')
    divergences: int | None = None
    decisions: int | None = None
    seconds: float = 0.0

    @property
    def row_id(self) -> str:
        base = f"{self.target.short}/{self.seed_set.short}"
        return base if len(GA_SEEDS) == 1 else f"{base}#{self.ga_seed}"

    @property
    def effective_cosmetic(self) -> float:
        """The penalty as it actually entered the fitness the GA maximised."""
        return self.cosmetic * self.cosmetic_scale


# ═════════════════════════════════════════════════════════════════════════════
# Runner
# ═════════════════════════════════════════════════════════════════════════════

def _apply_overrides(module) -> None:
    """Push the RUN SETTINGS overrides onto a driver module.

    The drivers read N_ROLLOUTS / HORIZON as module globals at call time and
    pass them into their `sr.TraceCache` per call, so setting them here changes
    what `module.conformance_counts` measures and invalidates any target trace
    recorded under the previous values.
    """
    for name, value in (("N_ROLLOUTS", N_ROLLOUTS), ("HORIZON", HORIZON)):
        if value is not None:
            setattr(module, name, value)


def _rows(spec: ExperimentSpec) -> list[tuple[Variant, SeedSet]]:
    if PAIRING == 'cross':
        return [(t, s) for t in spec.targets for s in spec.seed_sets]
    if len(spec.targets) != len(spec.seed_sets):
        raise ValueError(
            f"{spec.key}: PAIRING='zip' needs as many targets as seed sets "
            f"({len(spec.targets)} vs {len(spec.seed_sets)}); use PAIRING='cross'."
        )
    return list(zip(spec.targets, spec.seed_sets))


def run_row(spec: ExperimentSpec, target: Variant, seed_set: SeedSet,
            ga_seed: int) -> RowResult:
    """Run one GA against *target*, started from *seed_set*."""
    mod = spec.module

    def counts_fn(tree):
        return spec.counts(mod, tree, target.tree)

    def evaluate_primary(tree):
        return spec.primary(mod, counts_fn(tree))

    evaluate = sr.make_evaluator(evaluate_primary, mod.COSMETIC_CONFIG)

    experiment = sr.Experiment(
        name=f"{spec.title} | {target.short} {target.label} | "
             f"{seed_set.short} {seed_set.label}",
        reference_label=target.label,
        config=mod.CONFIG,
        evaluate=evaluate,
        evaluate_primary=evaluate_primary,
        cosmetic_config=mod.COSMETIC_CONFIG,
        seed_policies=list(seed_set.trees),
        seed_policy=seed_set.trees[0],
        pop_size=POP_SIZE or mod.POP_SIZE,
        n_generations=N_GENERATIONS or mod.N_GENERATIONS,
        tournament_k=mod.TOURNAMENT_K,
        p_crossover=mod.P_CROSSOVER,
        elite_k=mod.ELITE_K,
        reference_policy=target.tree,
        conformance_counts=counts_fn,
        scale_config=getattr(mod, "SCALE_CONFIG", None),
    )

    curve:  list[float] = []
    scales: list[float] = []

    def record(rec):
        curve.append(rec['best_fit'])
        scales.append(rec.get('cosmetic_scale', 1.0))

    t0 = time.time()
    best_tree, best_fit = sr.run_ga(
        experiment, seed=ga_seed, verbose=VERBOSE_GA, on_generation=record,
    )
    elapsed = time.time() - t0

    counts   = counts_fn(best_tree)
    primary  = evaluate_primary(best_tree)
    cosmetic = cosmetic_penalty(best_tree, mod.COSMETIC_CONFIG)

    # `best_fit` is scored at the last generation's cosmetic scale, so the
    # target/seed reference numbers have to be scored the same way or the three
    # are not on one axis (and the target line in the plot lands wrong).
    scale = scales[-1] if scales else 1.0
    evaluate_at_scale = sr.make_evaluator(evaluate_primary, mod.COSMETIC_CONFIG,
                                          scale)

    return RowResult(
        key=spec.key, target=target, seed_set=seed_set, ga_seed=ga_seed,
        fitness=best_fit,
        primary=float('nan') if primary is None else primary,
        cosmetic=cosmetic,
        discovered=best_tree,
        cosmetic_scale=scale,
        curve=curve,
        target_fitness=evaluate_at_scale(target.tree),
        seed_fitness=max(evaluate_at_scale(s) for s in seed_set.trees),
        divergences=None if counts is None else counts[0],
        decisions=None if counts is None else counts[1],
        seconds=elapsed,
    )


def run_experiment(spec: ExperimentSpec) -> list[RowResult]:
    _apply_overrides(spec.module)
    results = []
    for target, seed_set in _rows(spec):
        for ga_seed in GA_SEEDS:
            print(f"  [{spec.key}] {target.short} {target.label}  <-  "
                  f"{seed_set.short} {seed_set.label}  (ga_seed={ga_seed}) ...",
                  flush=True)
            res = run_row(spec, target, seed_set, ga_seed)
            rate = ("n/a" if not res.decisions
                    else f"{res.divergences / res.decisions:.3f}")
            print(f"      fitness={res.fitness:.4f}  "
                  f"(primary={res.primary:.4f}, "
                  f"cosmetic=-{res.effective_cosmetic:.4f} "
                  f"[raw {res.cosmetic:.4f} x{res.cosmetic_scale:.3g}])  "
                  f"divergence rate={rate}  [{res.seconds:.0f} s]", flush=True)
            print(f"      {repr(res.discovered)}", flush=True)
            results.append(res)
    return results


# ═════════════════════════════════════════════════════════════════════════════
# Reporting
# ═════════════════════════════════════════════════════════════════════════════

def _cell(short: str, label: str, trees) -> str:
    """A policy cell: id + description, then the policy itself, one per line."""
    if not isinstance(trees, (list, tuple)):
        trees = [trees]
    body = "\n".join(repr(t) for t in trees)
    return f"{short} - {label}\n{body}"


def table_for(results: Sequence[RowResult]) -> pd.DataFrame:
    return pd.DataFrame([{
        "target policy":    _cell(r.target.short, r.target.label, r.target.tree),
        "seed policy":      _cell(r.seed_set.short, r.seed_set.label,
                                  list(r.seed_set.trees)),
        "fitness":          r.fitness,
        "fitness primary":  r.primary,
        # The penalty as it entered the fitness, i.e. after the per-generation
        # rescaling, so `fitness = primary - cosmetic` still reads true here.
        # The raw penalty and the scale are broken out in `diagnostics`.
        "fitness cosmetic": r.effective_cosmetic,
        "discovered policy": repr(r.discovered),
    } for r in results])


def diagnostics_for(results: Sequence[RowResult]) -> pd.DataFrame:
    return pd.DataFrame([{
        "experiment":        r.key,
        "row":               r.row_id,
        "ga seed":           r.ga_seed,
        "target fitness":    r.target_fitness,
        "best seed fitness": r.seed_fitness,
        "best fitness":      r.fitness,
        "cosmetic raw":      r.cosmetic,
        "cosmetic scale":    r.cosmetic_scale,
        "divergences":       r.divergences,
        "decisions":         r.decisions,
        "divergence rate":   (None if not r.decisions
                              else r.divergences / r.decisions),
        "generations":       max(len(r.curve) - 1, 0),
        "runtime (s)":       round(r.seconds, 1),
    } for r in results])


def write_results_json(key: str, results: Sequence[RowResult],
                       spec: ExperimentSpec, out_dir: Path) -> Path:
    """Dump one experiment's payload to `results_<key>.json`.

    Written per experiment rather than once at the end: the full sweep takes
    the better part of an hour, and losing all of it to an interruption in the
    last driver is not a risk worth carrying.  It also lets the four drivers
    run as separate processes -- `report.py` merges whatever files it finds.
    """
    import json
    payload = results_json({key: list(results)}, {key: spec})
    path = _writable(out_dir / f"results_{key}.json")
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=1)
    return path


def results_json(all_results: dict[str, list[RowResult]],
                 specs: dict[str, ExperimentSpec]) -> dict:
    """Everything report.py needs, including the paper-style policy listings.

    The workbook carries one-line `repr`s, which is what a spreadsheet wants;
    the report wants the multi-line algorithm form, and it should not have to
    re-import the drivers to rebuild the trees just to render them.
    """
    from policy_listing import listing

    out: dict = {"runs": {}}
    for key, results in all_results.items():
        spec = specs[key]
        rows = []
        for r in results:
            # "Recovered" means behaviourally equivalent on the sample the
            # search saw -- zero divergences over every decision point reached
            # -- not textually identical to the target.  A smaller equivalent
            # spelling is a success, not a miss.
            recovered = r.divergences == 0 and bool(r.decisions)
            rows.append({
                "process":            spec.title.split(" (")[0],
                "title":              spec.title,
                "target_short":       r.target.short,
                "target_label":       r.target.label,
                "target_listing":     listing(r.target.tree),
                "seed_short":         r.seed_set.short,
                "seed_label":         r.seed_set.label,
                "seed_listings":      [listing(t) for t in r.seed_set.trees],
                "fitness":            r.fitness,
                "primary":            r.primary,
                "cosmetic_raw":       r.cosmetic,
                "cosmetic_effective": r.effective_cosmetic,
                "discovered_repr":    repr(r.discovered),
                "discovered_listing": listing(r.discovered),
                "target_fitness":     r.target_fitness,
                "seed_fitness":       r.seed_fitness,
                "divergences":        r.divergences,
                "decisions":          r.decisions,
                "recovered":          recovered,
                "seconds":            r.seconds,
                "curve":              r.curve,
                "ga_seed":            r.ga_seed,
            })
        out["runs"][key] = rows
    return out


def curves_frame(all_results: dict[str, list[RowResult]]) -> pd.DataFrame:
    rows = []
    for key, results in all_results.items():
        for r in results:
            for gen, fit in enumerate(r.curve):
                rows.append({
                    "experiment": key,
                    "row": r.row_id,
                    "target": r.target.label,
                    "seeds": r.seed_set.label,
                    "generation": gen,
                    "best fitness": fit,
                })
    return pd.DataFrame(rows)


def settings_frame() -> pd.DataFrame:
    return pd.DataFrame([
        ("run at",        datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("ga seeds",      ", ".join(str(s) for s in GA_SEEDS)),
        ("pairing",       PAIRING),
        ("experiments",   ", ".join(EXPERIMENTS)),
        ("pop size",      "driver default" if POP_SIZE is None else POP_SIZE),
        ("generations",   "driver default" if N_GENERATIONS is None else N_GENERATIONS),
        ("rollouts",      "driver default" if N_ROLLOUTS is None else N_ROLLOUTS),
        ("horizon",       "driver default" if HORIZON is None else HORIZON),
        ("fitness",       "fitness = fitness primary - fitness cosmetic"),
        ("cosmetic scale", "per-generation, bounded by the population's primary "
                           "spread (ga_fitness.ScaleConfig)"),
    ], columns=["setting", "value"])


def _writable(path: Path) -> Path:
    """`path`, or a timestamped sibling when `path` cannot be opened.

    The workbook is written *last*, after every GA has run.  If it happens to
    be open in Excel the write raises PermissionError and hours of simulation
    go in the bin -- so fall back to a new name rather than fail.
    """
    try:
        with open(path, "ab"):
            return path
    except OSError:
        alt = path.with_name(f"{path.stem}_{datetime.now():%Y%m%d_%H%M%S}{path.suffix}")
        print(f"  ! {path.name} is locked (open elsewhere?) -- writing {alt.name}")
        return alt


def write_excel(all_results: dict[str, list[RowResult]],
                specs: dict[str, ExperimentSpec], path: Path) -> Path:
    from openpyxl.styles import Alignment, Font

    path = _writable(path)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        settings_frame().to_excel(writer, sheet_name="settings", index=False)
        for key, results in all_results.items():
            table_for(results).to_excel(writer, sheet_name=specs[key].sheet,
                                        index=False)
        diagnostics = pd.concat([diagnostics_for(r) for r in all_results.values()],
                                ignore_index=True)
        diagnostics.to_excel(writer, sheet_name="diagnostics", index=False)

        widths = {"target policy": 60, "seed policy": 60, "discovered policy": 70,
                  "fitness": 12, "fitness primary": 15, "fitness cosmetic": 16,
                  "setting": 16, "value": 46}
        for sheet in writer.book.worksheets:
            headers = [c.value for c in sheet[1]]
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(vertical="top")
            for idx, header in enumerate(headers, start=1):
                letter = sheet.cell(row=1, column=idx).column_letter
                sheet.column_dimensions[letter].width = widths.get(header, 18)
                for row in range(2, sheet.max_row + 1):
                    c = sheet.cell(row=row, column=idx)
                    c.alignment = Alignment(vertical="top", wrap_text=True)
                    if isinstance(c.value, float):
                        c.number_format = "0.0000"
            sheet.freeze_panes = "A2"
    return path


def plot_experiment(key: str, results: Sequence[RowResult], title: str,
                    path: Path) -> None:
    """Best fitness over generations.

    Color is the one distinction these plots exist to show -- whether the run
    started with pre-knowledge -- so every row is colored by its seed set, not
    by its position in the table.  With several GA seeds per seed set this
    draws several same-colored lines rather than one, which is the point: the
    spread within a color is the run-to-run variance for that condition.
    """
    fig, ax = plt.subplots(figsize=FIG_SIZE, dpi=FIG_DPI)
    fig.patch.set_facecolor(SURFACE)
    ax.set_facecolor(SURFACE)

    for r in results:
        color = _preknowledge_color(r.seed_set.label)
        ax.plot(range(len(r.curve)), r.curve, color=color,
                linewidth=1.6, alpha=0.85, solid_capstyle="round")

    if results:
        # Every row in one plot targets the same policy, so its own achievable
        # fitness (zero divergences, minus its own cosmetic penalty) is one
        # reference line, not one per row.  Not strictly a ceiling -- a
        # smaller tree that is behaviourally equivalent scores *above* it.
        ax.axhline(results[0].target_fitness, color=TEXT_SECONDARY,
                   linewidth=1.0, linestyle=(0, (2, 3)), alpha=0.6, zorder=1)

    _style_axes(ax, title)
    _preknowledge_legend(ax)

    fig.tight_layout()
    fig.savefig(path, facecolor=SURFACE, bbox_inches="tight")
    plt.close(fig)


def plot_from_curves(curves: pd.DataFrame, specs: dict[str, ExperimentSpec],
                     out_dir: Path) -> None:
    """Re-draw the plots from a saved curves.csv (no GA run).

    Colored the same way as `plot_experiment`: by pre-knowledge, not by row,
    so several GA seeds of one seed set draw as several same-colored lines.
    """
    for key, group in curves.groupby("experiment", sort=False):
        fig, ax = plt.subplots(figsize=FIG_SIZE, dpi=FIG_DPI)
        fig.patch.set_facecolor(SURFACE)
        ax.set_facecolor(SURFACE)
        for row_id, run in group.groupby("row", sort=False):
            color = _preknowledge_color(run["seeds"].iloc[0])
            ax.plot(run["generation"], run["best fitness"], color=color,
                    linewidth=1.6, alpha=0.85)
        _style_axes(ax, specs[key].title if key in specs else key)
        _preknowledge_legend(ax)
        fig.tight_layout()
        fig.savefig(out_dir / f"{key}.png", facecolor=SURFACE,
                    bbox_inches="tight")
        plt.close(fig)


# ═════════════════════════════════════════════════════════════════════════════
# Experiment definitions -- one builder per driver
#
# Each builder names ONE target -- the normative policy the paper prints for
# that process, which is the driver's own REFERENCE_POLICY -- and TWO seed
# sets: S0 with no pre-knowledge (policies drawn at random from the grammar)
# and S1 with some pre-knowledge (the driver's hand-built partial seeds).
# Under PAIRING='cross' that is the two rows T/S0 and T/S1, so the four
# drivers together give the eight runs.
# ═════════════════════════════════════════════════════════════════════════════


def _uninformed(mod, rng_seed: int = 7):
    """One policy drawn at random from the driver's own grammar.

    The "no pre-knowledge" start of Section 5: a seed policy that is
    *initialized randomly* rather than derived from expert knowledge.  Drawn
    with a fixed RNG so the run is reproducible and the drawn tree can be
    reported next to the informed one.

    One, not several, to match Algorithm~
ef{alg:ga-init}, which initialises
    the population from a single seed policy pi_0.
    """
    import random as _random
    rng = _random.Random(rng_seed)
    return [sr.rnd_policy(rng, mod.CONFIG)]


def _spec_batching() -> ExperimentSpec:
    import ga_batching_conf as mod

    return ExperimentSpec(
        key="batching",
        module=mod,
        title="Cat I - Transportation, simultaneous execution (pi_batch)",
        targets=[
            Variant("T", f"pi_batch (Alg. Cat I, batch > {mod.BATCH_SIZE})",
                    mod.REFERENCE_POLICY),
        ],
        seed_sets=[
            SeedSet("S0", "no pre-knowledge (random)",
                    _uninformed(mod, rng_seed=11)),
            SeedSet("S1", "some pre-knowledge", [mod.SEED_POLICY]),
        ],
        counts=lambda m, tree, ref: m.conformance_counts(tree, ref),
        primary=lambda m, counts: sr.primary_rate(counts),
    )


def _spec_choice_sl() -> ExperimentSpec:
    import ga_choice_SL_conf as mod

    return ExperimentSpec(
        key="choice_sl",
        module=mod,
        title="Cat II-a - Assembly, time-based SLA (pi_sla)",
        targets=[
            Variant("T", "pi_sla (Alg. Cat II-a, phone first, SLA 0.95)",
                    mod.REFERENCE_POLICY),
        ],
        seed_sets=[
            SeedSet("S0", "no pre-knowledge (random)",
                    _uninformed(mod, rng_seed=12)),
            SeedSet("S1", "some pre-knowledge", [mod.SEED_POLICY]),
        ],
        counts=lambda m, tree, ref: m.conformance_counts(tree, ref),
        primary=lambda m, counts: sr.primary_rate(counts),
    )


def _spec_ratelimit() -> ExperimentSpec:
    import ga_ratelimit_conf as mod

    return ExperimentSpec(
        key="ratelimit",
        module=mod,
        title="Cat II-b - Dual manufacturing, number of executions (pi_rate)",
        targets=[
            Variant("T", f"pi_rate (Alg. Cat II-b, caps "
                         f"{mod.MAX_PER_TIME_A}/{mod.MAX_PER_TIME_B})",
                    mod.REFERENCE_POLICY),
        ],
        seed_sets=[
            SeedSet("S0", "no pre-knowledge (random)",
                    _uninformed(mod, rng_seed=13)),
            SeedSet("S1", "some pre-knowledge", [mod.SEED_POLICY]),
        ],
        counts=lambda m, tree, ref: m.conformance_counts(tree, ref),
        primary=lambda m, counts: sr.primary_rate(counts),
    )


def _spec_priority() -> ExperimentSpec:
    import ga_priority_conf as mod

    return ExperimentSpec(
        key="priority",
        module=mod,
        title="Cat II-c - Manufacturing, priority on data (pi_prio)",
        targets=[
            Variant("T", "pi_prio (Alg. Cat II-c, MAX vs MAX)",
                    mod.REFERENCE_POLICY),
        ],
        seed_sets=[
            SeedSet("S0", "no pre-knowledge (random)",
                    _uninformed(mod, rng_seed=14)),
            SeedSet("S1", "some pre-knowledge", [mod.SEED_POLICY]),
        ],
        counts=lambda m, tree, ref: m.conformance_counts(tree, ref),
        primary=lambda m, counts: sr.primary_rate(counts),
    )


SPEC_BUILDERS = {
    "batching":  _spec_batching,   # Cat I
    "choice_sl": _spec_choice_sl,  # Cat II-a
    "ratelimit": _spec_ratelimit,  # Cat II-b
    "priority":  _spec_priority,   # Cat II-c
}


# ═════════════════════════════════════════════════════════════════════════════
# Entry point
# ═════════════════════════════════════════════════════════════════════════════

def main(argv: Sequence[str]) -> int:
    keys = [a for a in argv if not a.startswith("-")] or list(EXPERIMENTS)
    unknown = [k for k in keys if k not in SPEC_BUILDERS]
    if unknown:
        print(f"unknown experiment(s): {', '.join(unknown)}\n"
              f"available: {', '.join(SPEC_BUILDERS)}")
        return 2

    OUT_DIR.mkdir(parents=True, exist_ok=True)

    if "--list" in argv:
        for key in keys:
            spec = SPEC_BUILDERS[key]()
            print(f"\n{key}  ({spec.module.__name__})")
            for target, seed_set in _rows(spec):
                print(f"  {target.short}/{seed_set.short}  "
                      f"target: {target.label}  |  seeds: {seed_set.label}")
        return 0

    specs = {key: SPEC_BUILDERS[key]() for key in keys}

    if "--plot-only" in argv:
        curves = pd.read_csv(OUT_DIR / "curves.csv")
        plot_from_curves(curves[curves["experiment"].isin(keys)], specs, OUT_DIR)
        print(f"plots rewritten in {OUT_DIR}")
        return 0

    started = time.time()
    all_results: dict[str, list[RowResult]] = {}
    for key in keys:
        spec = specs[key]
        print(f"\n=== {spec.title}  ({spec.module.__name__}) ===", flush=True)
        all_results[key] = run_experiment(spec)
        plot_experiment(key, all_results[key], spec.title, OUT_DIR / f"{key}.png")
        print(f"  -> {write_results_json(key, all_results[key], spec, OUT_DIR)}",
              flush=True)

    xlsx = write_excel(all_results, specs, OUT_DIR / XLSX_NAME)
    curves = _writable(OUT_DIR / "curves.csv")
    curves_frame(all_results).to_csv(curves, index=False)

    import json
    payload = _writable(OUT_DIR / "results.json")
    with open(payload, "w", encoding="utf-8") as fh:
        json.dump(results_json(all_results, specs), fh, indent=1)

    print(f"\nDone in {(time.time() - started) / 60:.1f} min.")
    print(f"  {xlsx}")
    print(f"  {curves}")
    print(f"  {payload}")
    for key in keys:
        print(f"  {OUT_DIR / (key + '.png')}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
